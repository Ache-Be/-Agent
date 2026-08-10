"""
随堂测验（xlsx）数据解析器。

文件名格式：项目0-基本数据类型.xlsx（项目编号-章节名）

Sheet 结构（由智慧树平台导出）：
  - 班级参与详情：课程/班级/活动名称/随堂测试总分/平均分
  - 学生得分与正确率：学生姓名/学号/班级/参与答题数目/得分/正确率
  - 试题详情：题干/题目类型/正确答案/题目分值/答对人数/答错人数/答题正确率
  - 第一题/第二题/...：每道题的学生作答明细
"""

import re
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import load_workbook

from analysis.config import load_config


def parse_quiz_filename(filename: str) -> Dict:
    """从随堂测验文件名解析项目编号和章节名。

    "项目0-基本数据类型.xlsx" → {"project_id": "0", "chapter_name": "基本数据类型"}
    """
    name = Path(filename).stem
    m = re.match(r"项目\s*(\d+)\s*[-—–_]\s*(.+)", name)
    if m:
        return {"project_id": m.group(1), "chapter_name": m.group(2).strip()}
    return {"project_id": "", "chapter_name": name}


def parse_quiz_file(filepath: str) -> Dict:
    """解析单个随堂测验 xlsx，返回结构化数据。"""
    wb = load_workbook(filepath, data_only=True)
    file_info = parse_quiz_filename(Path(filepath).name)

    result = {
        "file_path": filepath,
        "project_id": file_info["project_id"],
        "chapter_name": file_info["chapter_name"],
        "class_name": "",
        "activity_name": "",
        "total_score": 0,
        "avg_score": 0,
        "student_count": 0,
        "students": [],
        "questions": [],
    }

    # ---- Sheet: 班级参与详情 ----
    if "班级参与详情" in wb.sheetnames:
        rows = list(wb["班级参与详情"].iter_rows(values_only=True))
        if len(rows) > 1:
            header = [_cell_str(c) for c in rows[0]]
            row_dict = _row_to_dict(header, rows[1])
            result["class_name"] = row_dict.get("班级", "")
            result["activity_name"] = row_dict.get("活动名称", "")
            result["total_score"] = _safe_float(row_dict.get("随堂测试总分"))
            result["avg_score"] = _safe_float(row_dict.get("平均分"))

    # ---- Sheet: 学生得分与正确率 ----
    if "学生得分与正确率" in wb.sheetnames:
        rows = list(wb["学生得分与正确率"].iter_rows(values_only=True))
        if rows:
            header = [_cell_str(c) for c in rows[0]]
            for row in rows[1:]:
                if not row or not any(c is not None for c in row):
                    continue
                row_dict = _row_to_dict(header, row)
                name = row_dict.get("学生姓名", "")
                sid = row_dict.get("学号", "")
                if not name and not sid:
                    continue
                result["students"].append({
                    "name": name,
                    "student_id": sid,
                    "class_name": row_dict.get("班级", ""),
                    "answer_count": _safe_int(row_dict.get("参与答题数目")),
                    "score": _safe_float(row_dict.get("得分")),
                    "accuracy": _parse_percent(row_dict.get("正确率")),
                })
        result["student_count"] = len(result["students"])

    # ---- Sheet: 试题详情 ----
    if "试题详情" in wb.sheetnames:
        rows = list(wb["试题详情"].iter_rows(values_only=True))
        if rows:
            header = [_cell_str(c) for c in rows[0]]
            for row in rows[1:]:
                if not row or not any(c is not None for c in row):
                    continue
                row_dict = _row_to_dict(header, row)
                result["questions"].append({
                    "question": row_dict.get("题干", ""),
                    "qtype": row_dict.get("题目类型", ""),
                    "answer": row_dict.get("正确答案", ""),
                    "score": _safe_float(row_dict.get("题目分值")),
                    "correct_rate": _parse_percent(row_dict.get("答题正确率")),
                })

    wb.close()
    return result


def scan_quiz_directory(root_dir: str) -> List[Dict]:
    """递归扫描随堂测验目录下的所有 xlsx 文件。"""
    results = []
    root = Path(root_dir)
    if not root.exists():
        return results
    for xlsx in sorted(root.rglob("*.xlsx")):
        results.append({
            "class_dir": xlsx.parent.name,
            "file_path": str(xlsx),
            "filename": xlsx.name,
        })
    return results


def analyze_quiz_file(filepath: str, weak_threshold: Optional[float] = None) -> Dict:
    """分析单个随堂测验：统计薄弱学生（正确率低于阈值）。

    非随堂测验文件（如成绩导出表，无学生数据）会抛出 ValueError。
    """
    if weak_threshold is None:
        weak_threshold = load_config()["weak_threshold"]
    quiz = parse_quiz_file(filepath)
    if quiz["student_count"] == 0:
        raise ValueError("未识别到学生答题数据，可能不是随堂测验导出文件")
    students = quiz["students"]
    total = len(students)
    accs = [s["accuracy"] for s in students if s["accuracy"] is not None]
    weak = [s for s in students if s["accuracy"] is not None and s["accuracy"] < weak_threshold]
    quiz["weak_count"] = len(weak)
    quiz["weak_rate"] = round(len(weak) / total, 2) if total else 0
    quiz["avg_accuracy"] = round(sum(accs) / len(accs), 2) if accs else 0
    return quiz


def analyze_quiz_directory(root_dir: str) -> List[Dict]:
    """批量分析随堂测验目录，返回所有测验结果。"""
    results = []
    for f in scan_quiz_directory(root_dir):
        try:
            r = analyze_quiz_file(f["file_path"])
            r["class_dir"] = f["class_dir"]
            results.append(r)
        except Exception as e:
            results.append({
                "file_path": f["file_path"],
                "filename": f["filename"],
                "error": str(e),
            })
    return results


# ==================== 工具函数 ====================

def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _row_to_dict(header: List[str], row) -> Dict:
    d = {}
    for i, h in enumerate(header):
        if not h:
            continue
        v = row[i] if i < len(row) else None
        d[h] = _cell_str(v)
    return d


def _safe_float(v) -> Optional[float]:
    if v is None or v == "" or v == "--":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _safe_int(v) -> Optional[int]:
    f = _safe_float(v)
    return int(f) if f is not None else None


def _parse_percent(v) -> Optional[float]:
    """"100.00%" → 1.0；"85.71%" → 0.8571；纯数字按百分数处理"""
    if v is None or v == "" or v == "--":
        return None
    s = str(v).strip()
    if s.endswith("%"):
        try:
            return round(float(s[:-1]) / 100.0, 4)
        except (ValueError, TypeError):
            return None
    f = _safe_float(s)
    if f is not None and f > 1:
        return round(f / 100.0, 4)
    return f
