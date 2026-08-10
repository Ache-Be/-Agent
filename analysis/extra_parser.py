"""
单元练习 / 课堂活动分数明细 解析。

这两类文件都是智慧树导出的 xlsx，与随堂测验格式不同：
  1. 单元练习（如「苏理工软件(单元练习).xlsx」）：含 作业/考试/课堂活动 三个 Sheet
     - 作业：多个单元练习（每单元 成绩+状态 两列一组）
     - 考试：期末综合测试（成绩+状态）
     - 课堂活动：课程活动总分
  2. 课堂活动分数明细（如「课堂活动分数明细-苏理工软件.xlsx」）：含 班级数据概览/学生数据概览/学生签到数据明细
     - 学生数据概览：签到、随堂测试作答/正确率、点名/抢答/投票/问卷/头脑风暴、课堂活动总分

统一输出结构（与 quiz_parser 风格一致）：
  {
    "type": "unit" | "attendance",
    "class_name": "...",
    "student_count": N,
    "students": [
      {
        "name": "...", "student_id": "...", "class_name": "...",
        # unit 类型：
        "units": [{"name": "第一单元练习", "score": 130.0, "max": 130.0, "status": "已批阅"}],
        "exam": {"name": "期末综合测试", "score": 260.0, "max": 275.0, "status": "已完成考试"},
        "activity_score": 40.56,
        # attendance 类型：
        "signed_ratio": 1.0, "test_accuracy": 0.83, "activity_total": 245.0, ...
      }, ...
    ]
  }
"""

import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


def _parse_percent(s) -> Optional[float]:
    """'100.00%' → 1.0；纯数字原样返回；失败返回 None"""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    m = re.match(r"^([\d.]+)%$", s)
    if m:
        return float(m.group(1)) / 100.0
    try:
        return float(s)
    except ValueError:
        return None


def _safe_float(s) -> Optional[float]:
    if s is None:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _find_header_row(rows, keyword="学生姓名", limit=10):
    """在行列表中查找包含关键词的表头行索引"""
    for i, r in enumerate(rows[:limit]):
        if r and r[0] and keyword in str(r[0]):
            return i
    return -1


def _parse_max(full: str) -> Optional[float]:
    """'成绩（满分：130）' → 130.0"""
    m = re.search(r"满分[:：]\s*([\d.]+)", str(full or ""))
    return float(m.group(1)) if m else None


def _parse_two_level_sheet(ws, student_map: Dict, class_name: str):
    """解析两级表头 Sheet（作业/考试）：科目列 + 成绩/状态 两列一组。

    表头：
      行r:   学生姓名 | 学生学号 | 自定义标签 | 第一单元练习 | '' | 第二单元练习 | '' | ...
      行r+1: ''      | ''      | ''          | 成绩（满分：130） | 状态 | 成绩（满分：75） | 状态 | ...
    """
    rows = list(ws.iter_rows(values_only=True))
    hr = _find_header_row(rows)
    if hr < 0 or hr + 1 >= len(rows):
        return
    header = rows[hr]
    sub = rows[hr + 1]

    # 收集科目组：(科目名, 成绩列, 状态列)
    groups = []
    col = 3
    while col < len(header):
        name = str(header[col] or "").strip()
        score_label = str(sub[col] or "").strip() if col < len(sub) else ""
        status_label = str(sub[col + 1] or "").strip() if col + 1 < len(sub) else ""
        if not name:
            col += 1
            continue
        groups.append((name, col, col + 1, score_label, status_label))
        col += 2

    for r in rows[hr + 2:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        sid = str(r[1] or "").strip()
        if not name or not sid:
            continue
        stu = student_map.setdefault(sid, {"name": name, "student_id": sid, "class_name": class_name,
                                           "units": [], "exam": None, "activity_score": None})
        stu["name"] = name
        for gname, scol, stcol, score_label, status_label in groups:
            raw_score = r[scol] if scol < len(r) else None
            status = str(r[stcol] or "").strip() if stcol < len(r) else ""
            score = _safe_float(raw_score)
            if gname == "期末综合测试":
                max_score = _parse_max(score_label) or 0
                stu["exam"] = {"name": gname, "score": score or 0, "max": max_score, "status": status}
            else:
                max_score = _parse_max(score_label) or 0
                stu["units"].append({
                    "name": gname,
                    "score": score or 0,
                    "max": max_score,
                    "status": status or "已批阅",
                })


def _parse_activity_sheet(ws, student_map: Dict, class_name: str):
    """解析单级表头 Sheet（课堂活动）：学生姓名 | 学生学号 | 自定义标签 | 课程活动总分"""
    rows = list(ws.iter_rows(values_only=True))
    hr = _find_header_row(rows)
    if hr < 0:
        return
    score_col = 3
    for r in rows[hr + 1:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        sid = str(r[1] or "").strip()
        if not name or not sid:
            continue
        stu = student_map.setdefault(sid, {"name": name, "student_id": sid, "class_name": class_name,
                                           "units": [], "exam": None, "activity_score": None})
        stu["name"] = name
        stu["activity_score"] = _safe_float(r[score_col]) if score_col < len(r) else None


def analyze_unit_file(path) -> Dict:
    """解析单元练习 xlsx"""
    wb = load_workbook(path, data_only=True)
    class_name = Path(path).stem
    student_map = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn in ("作业", "考试"):
            _parse_two_level_sheet(ws, student_map, class_name)
        elif sn == "课堂活动":
            _parse_activity_sheet(ws, student_map, class_name)
    wb.close()

    if not student_map:
        raise ValueError("未解析到单元练习学生数据（不是单元练习文件？）")

    students = list(student_map.values())
    # 班级名优先取数据中出现最多的
    from collections import Counter
    cnt = Counter(s["class_name"] for s in students)
    real_class = cnt.most_common(1)[0][0] if cnt else class_name
    for s in students:
        s["class_name"] = real_class

    return {
        "type": "unit",
        "class_name": real_class,
        "student_count": len(students),
        "students": students,
    }


# ---- 课堂活动分数明细 ----

ATTEND_FIELDS = {
    4: "publish_count",     # 发布签到次数
    5: "sign_count",        # 签到次数
    6: "late",              # 迟到
    7: "early",             # 早退
    11: "signed_ratio",     # 签到出勤率*
    12: "test_published",   # 发布测试题数量
    13: "test_answered",    # 学生答题数量
    14: "test_ratio",       # 学生作答率
    16: "test_accuracy",    # 学生正确率**
    29: "activity_total",   # 课堂活动总分
}


def analyze_attendance_file(path) -> Dict:
    """解析课堂活动分数明细 xlsx（学生数据概览 Sheet）"""
    wb = load_workbook(path, data_only=True)
    if "学生数据概览" not in wb.sheetnames:
        wb.close()
        raise ValueError("缺少『学生数据概览』Sheet（不是课堂活动分数明细文件？）")
    ws = wb["学生数据概览"]
    rows = list(ws.iter_rows(values_only=True))
    # 行0=大类 行1=字段名 行2+=数据
    if not rows or len(rows) < 3:
        wb.close()
        raise ValueError("学生数据概览为空")

    students = []
    for r in rows[2:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        sid = str(r[1] or "").strip()
        if not name or not sid:
            continue
        stu = {
            "name": name,
            "student_id": sid,
            "class_name": str(r[3] or "").strip(),
        }
        for col, key in ATTEND_FIELDS.items():
            val = r[col] if col < len(r) else None
            if key in ("signed_ratio", "test_ratio", "test_accuracy"):
                stu[key] = _parse_percent(val)
            else:
                stu[key] = _safe_float(val)
        students.append(stu)
    wb.close()

    if not students:
        raise ValueError("未解析到课堂活动学生数据")

    class_name = students[0]["class_name"] or Path(path).stem
    return {
        "type": "attendance",
        "class_name": class_name,
        "student_count": len(students),
        "students": students,
    }


# ---- 每个知识点的学生学习数据完成情况 ----

def _strip_suffix(name: str) -> str:
    """去掉列名末尾的『学习进度』『掌握度』等后缀，还原知识点名"""
    for suf in ("学习进度", "掌握度", "完成情况"):
        if name.endswith(suf):
            return name[: -len(suf)].strip()
    return name.strip()


def analyze_knowledge_file(path) -> Dict:
    """解析智慧树导出的『每个知识点的学生学习数据完成情况.xlsx』。

    表头（sheet1）：
      班级 | 学号 | 姓名 | 自定义标签 | 总体掌握度 | 总体学习进度 | 已学/应学知识点
      | 知识点A 学习进度 | 知识点A 掌握度 | 知识点B 学习进度 | 知识点B 掌握度 | ...
    每个知识点两列一组（学习进度、掌握度）。

    输出结构（与 quiz_parser 风格一致）：
      {
        "type": "knowledge",
        "class_name": "...",
        "student_count": N,
        "students": [
          {
            "name": "...", "student_id": "...", "class_name": "...",
            "overall_mastery": 0.19,      # 总体掌握度（0~1，缺失为 None）
            "overall_progress": 0.377,    # 总体学习进度（0~1）
            "learned": "63/118",          # 已学/应学知识点
            "knowledge": [{"name": "...", "progress": 1.0, "mastery": 0.0}, ...],
          }, ...
        ]
      }
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("知识点完成情况文件为空")
    header = rows[0]
    if not header or str(header[4] or "") != "总体掌握度":
        raise ValueError("缺少『总体掌握度』列（不是知识点完成情况文件？）")

    # 收集知识点列：从第 8 列起两列一组（学习进度、掌握度）
    kp_groups = []
    col = 7
    while col + 1 < len(header):
        name = _strip_suffix(str(header[col] or ""))
        if not name:
            break
        kp_groups.append((name, col, col + 1))
        col += 2

    students = []
    for r in rows[1:]:
        if not r or not r[2]:
            continue
        name = str(r[2]).strip()
        sid = str(r[1] or "").strip()
        if not name and not sid:
            continue
        stu = {
            "name": name or sid,
            "student_id": sid,
            "class_name": str(r[0] or "").strip(),
            "overall_mastery": _parse_percent(r[4]) if len(r) > 4 else None,
            "overall_progress": _parse_percent(r[5]) if len(r) > 5 else None,
            "learned": str(r[6] or "").strip() if len(r) > 6 else "",
            "knowledge": [],
        }
        for kname, pcol, mcol in kp_groups:
            progress = _parse_percent(r[pcol]) if pcol < len(r) else None
            mastery = _parse_percent(r[mcol]) if mcol < len(r) else None
            if progress is not None or mastery is not None:
                stu["knowledge"].append({"name": kname, "progress": progress, "mastery": mastery})
        students.append(stu)

    if not students:
        raise ValueError("未解析到知识点完成情况学生数据")

    # 班级名取出现最多的
    from collections import Counter
    cnt = Counter(s["class_name"] for s in students if s["class_name"])
    class_name = cnt.most_common(1)[0][0] if cnt else Path(path).stem

    return {
        "type": "knowledge",
        "class_name": class_name,
        "student_count": len(students),
        "students": students,
    }
